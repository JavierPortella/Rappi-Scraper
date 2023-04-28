# Librerías a usar
from concurrent import futures as futures
from datetime import datetime, timedelta
import json
import logging
import os
import random
import re
import time
from traceback import TracebackException

from dotenv import dotenv_values
from openpyxl import load_workbook, Workbook
import pandas as pd
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from seleniumwire.utils import decode
from seleniumwire.webdriver import Chrome, ChromeOptions
from webdriver_manager.chrome import ChromeDriverManager

# Constantes
CURRENT_DATE = datetime.now().date()
ROOT_PATH = os.getcwd()
LOGGER = logging.getLogger(__name__)
THREAD = futures.ThreadPoolExecutor()
ENV = dotenv_values()
DATA_FILENAME = ENV["DATA_FILENAME"]
DATA_FOLDER = ENV["DATA_FOLDER"]
LOG_FILENAME = ENV["LOG_FILENAME"]
LOG_FOLDER = ENV["LOG_FOLDER"]
METADATA_FILENAME = ENV["METADATA_FILENAME"]
METADATA_SHEET_NAME = ENV["METADATA_SHEET_NAME"]
FB_USERNAME = ENV["FB_USERNAME"]
FB_PASSWORD = ENV["FB_PASSWORD"]



class Metadata:
    """Representa a la información generada durante la ejecución del scraper

    Attributes:
        start_time (float): Hora de inicio de la ejecución del scraper en segundos
        end_time (float): Hora de fin de la ejecución del scraper en segundos
        execution_date (str): Fecha de extracción de las categorias en formato %d/%m/%Y
        start_hour (str): Hora de inicio de la ejecución del scraper en formato %H:%M:%S
        end_hour (str): Hora de término de la ejecución del scraper en formato %H:%M:%S
        quantity (int): Cantidad de los productos extraídos de la página de Rappi
        time_execution (str): Tiempo de ejecución del scraper en segundos
        products_per_min (float): Cantidad de categorías que puede extraer el scraper en un minuto
        num_errors (int): Cantidad de errores ocurridos durante la ejecución del scraper
    """

    def __init__(self):
        """Genera todos los atributos para una instancia de la clase Metadata"""
        self._start_time = time.time()
        self._end_time = self._start_time
        self._execution_date = CURRENT_DATE.strftime("%d/%m/%Y")
        self._start_hour = time.strftime("%H:%M:%S", time.localtime(self._start_time))
        self._end_hour = self._start_hour
        self._quantity = 0
        self._time_execution = "0"
        self._products_per_min = 0
        self._num_errors = 0

    @property
    def num_errors(self):
        """Retorna el valor actual o actualiza el valor del atributo num_errors"""
        return self._num_errors

    @property
    def quantity(self):
        """Retorna el valor actual o actualiza el valor del atributo quantity"""
        return self._quantity

    @num_errors.setter
    def num_errors(self, num_errors):
        self._num_errors = num_errors

    @quantity.setter
    def quantity(self, quantity):
        self._quantity = quantity

    def set_attributes_values(self):
        """Establece los parámetros finales cuando se termina de ejecutar el scraper"""
        self._end_time = time.time()
        self._end_hour = time.strftime("%H:%M:%S", time.localtime(self._end_time))
        time_execution = self._end_time - self._start_time
        if time_execution > 0:
            self._time_execution = str(timedelta(seconds=time_execution)).split(".")[0]
            self._products_per_min = round((self._quantity * 60) / time_execution, 2)

    def print_metadata_information(self):
        """Imprime la información del tiempo de ejecución del scraper por consola"""
        LOGGER.info(f"Hora inicio: {self._start_hour}")
        LOGGER.info(f"Hora Fin: {self._end_hour}")
        LOGGER.info(f"Duración: {self._time_execution}")
        LOGGER.info(f"Productos Extraídos: {self._quantity}")
        LOGGER.info(f"Productos Extraídos / min: {self._products_per_min}")
        LOGGER.info(f"Número de errores: {self._num_errors}")


class Error(TracebackException):
    """Extiende la clase TracebackException para el manejo del traceback"""

    def __init__(self, error) -> None:
        """Genera todos los atributos para una instancia de la clase Error

        Args:
            error (Exception): Error ocurrido durante la ejecución del scraper
        """
        super().__init__(type(error), error, error.__traceback__)

    def print_error_detail(self):
        """Imprime toda la información del error por consola"""
        LOGGER.error("Ha ocurrido un error:")
        for line in self.format(chain=True):
            LOGGER.error(line)


class ScraperRappiProducts:
    """Representa a un bot para hacer web scraping en Rappi

    Attributes:
        metadata (Metadata): Objeto que maneja toda la información generada por el scraper durante su ejecución
        products (list): Lista que contiene todos los productos que ofertan los restaurantes en Rappi
        dataset (DataFrame): DataFrame que contiene toda la información extraída por el scraper
        links_to_go (list): Lista de restaurantes que en un primer intento no se pudo extraer su información
        driver (WebDriver): Objeto que maneja el navegador web
        wait (WebDriverWait): Objeto que maneja los tiempos de espera de búsqueda de elementos en la web
        action (ActionChains): Objeto que maneja las acciones que se pueden aplicar a los elementos en la web
    """

    def __init__(self):
        """Genera todos los atributos para una instancia de la clase ScraperRappiProducts"""
        LOGGER.info("Inicializando scraper")
        self._metadata = Metadata()
        self._products = []
        self._restaurants = []
        self._dataset = pd.DataFrame()
        self._links_to_go = []
        chrome_options = ChromeOptions()
        prefs = {
            "profile.default_content_setting_values.notifications": 2,
            "profile.managed_default_content_settings.popups": 2,
        }
        chrome_options.add_experimental_option("prefs", prefs)
        chrome_options.add_experimental_option(
            "excludeSwitches", ["enable-logging"]
        )  # Suprimir los mensajes de consola
        self._driver = Chrome(
            options=chrome_options,
            service=Service(ChromeDriverManager().install()),
        )
        self._driver.maximize_window()
        self._wait = WebDriverWait(self._driver, 10)
        self._action = ActionChains(self._driver)
        LOGGER.info("Scraper inicializado satisfactoriamente")

    def login(self, user_name, user_password):
        """Inicia sesión en la página web de Rapi usando una cuenta de Facebook

        Args:
            user_name (str): Usuario activo de facebook
            user_password (str): Contraseña del usuario activo de facebook
        """
        LOGGER.info("Iniciando sesión")
        self._driver.get("https://www.rappi.com.pe/login")
        # Usar la opción de facebook
        self._driver.find_element(
            By.XPATH, "//button[@class='chakra-button css-1hdh3ss']"
        ).click()
        time.sleep(random.uniform(3.0, 4.0))
        # Cambiar a la pestaña de inicio de sesión de Facebook
        self._driver.switch_to.window(self._driver.window_handles[1])
        # Completar los campos de usuario y contraseña
        username = self._wait.until(EC.presence_of_element_located((By.ID, "email")))
        password = self._wait.until(EC.presence_of_element_located((By.ID, "pass")))
        username.clear()
        password.clear()
        username.send_keys(user_name)
        password.send_keys(user_password)
        # Iniciar sesión
        self._driver.find_element(By.CSS_SELECTOR, "input[name='login']").click()
        # Dando permisos a Rappi en el primer inicio de sesión
        try:
            self._wait.until(
                EC.presence_of_element_located(
                    (
                        By.XPATH,
                        "//div[@class='x1r8uery x1iyjqo2']",
                    )
                )
            ).click()
        except:
            pass
        # Volver a la pestaña principal
        self._driver.switch_to.window(self._driver.window_handles[0])
        time.sleep(random.uniform(8.0, 10.0))
        # Detectar si el ícono del usuario aparece en la página
        self._wait.until(
            EC.presence_of_element_located(
                (By.XPATH, "//div[@class='sc-fdt2fy-11 dJNVzE']")
            )
        )
        LOGGER.info("Se inició sesión correctamente")

    def scrap_product(self, product, rest, status, cat):
        """Extrae la información de un producto de un restaurante

        Args:
            product (WebElement): Elemento web que contiene la información del producto de un restaurante
            rest (str): Nombre Restaurante
            status (str): Estado del restaurante
            cat (str): Categoría del restaurante

        Returns:
            list: Arreglo con toda la información del producto de un restaurante
        """
        data = []
        try:
            # Popularidad
            try:
                product.find_element(
                    By.XPATH, ".//p[@class='chakra-text css-n0gvg7']"
                ).text
                data.append(True)
            except:
                data.append(False)
            # Nombre
            data.append(
                product.find_element(By.XPATH, ".//div[@class='css-k008qs']").text
            )
            # Descripción
            data.append(
                product.find_element(
                    By.XPATH,
                    ".//p[@class='chakra-text sc-a04fe063-2 gHQcCO css-1rmjo0r']",
                ).text
            )
            # Precios
            prices = product.find_element(
                By.XPATH, ".//.//div[contains(@class, 'chakra-skeleton')]"
            ).text.split("S/ ")
            # Precio con descuento
            try:
                data.append(prices[-2])
            except:
                data.append(None)
            # Precio sin descuento
            data.append(prices[-1])
            # Nombre del restaurante
            data.append(rest)
            # Disponibilidad
            data.append(status)
            # Categoría
            data.append(cat)
            return data
        except Exception as error:
            try:
                LOGGER.error(f"Error al scrapear el producto {data[1]}")
            except:
                pass
            LOGGER.error(f"Error de tipo: {error.__class__}")
            return []

    def scrap_restaurante(self, rest, link, status):
        """Ingresa a un restaurante y extrae la información de todos los productos que ofrece

        Args:
            rest (str): Nombre del restaurante
            link (str): Link del restaurante
            status (str): Estado actual del restaurante
        """
        try:
            LOGGER.info(f"Extrayendo información del restaurante {rest}")
            self._driver.get(link)
            # Productos que ofrece el restaurante
            products = self._driver.find_elements(
                By.XPATH, '//div[@class="chakra-stack css-46p1lt"]'
            )
            # Categoría del restaurante
            try:
                category = self._driver.find_element(
                    By.XPATH, "//div[@class='sc-3627ee44-1 dEYlRK']/h2[2]"
                ).text
            except:
                LOGGER.info(f"El restaurante no posee ninguna categoría")
                category = None
            # Extrayendo la información de varios productos a la vez
            future_products = [
                THREAD.submit(self.scrap_product, product, rest, status, category)
                for product in products
            ]
            for future_product in futures.wait(future_products).done:
                self._products.append(future_product.result())
        except Exception as error:
            LOGGER.error("Fallo al extraer la información de los productos")
            LOGGER.error(f"Error de tipo: {error.__class__}")

    def extract_products(self, restaurant):
        """Extrae la información de los productos de un restaurante usando los recursos de red

        Args:
            restaurant (WebElement): Elemento web que representa a un restaurante
        """
        try:
            self._action.scroll_to_element(restaurant).move_to_element(
                restaurant
            ).perform()
            # Disponibilidad del restaurante
            try:
                restaurant_status = restaurant.find_element(
                    By.XPATH, ".//p[contains(@class,'chakra-text')]"
                ).text
            except:
                restaurant_status = "Restaurante abierto"
            # Enlace web del restaurante
            restaurant_link = restaurant.get_attribute("href")
            self._restaurants.append(restaurant_link)
            # Request que contiene la información del restaurante
            request = self._driver.wait_for_request(
                re.search("restaurantes/(.*)", restaurant_link).group(1),
                timeout=3,
            )
            # Obteniendo la respuesta decodificada del request
            decoded_body = decode(
                request.response.body,
                request.response.headers.get("Content-Encoding", "identity"),
            ).decode("utf-8")
            # Convirtiendo en formato json
            json_data = json.loads(decoded_body)
            rest_dict = json_data["pageProps"]["fallback"]
            restaurant_data = rest_dict[next(iter(rest_dict))]
            # Nombre y categoría del restaurante
            restaurant_name = restaurant_data.get("brandName")
            restaurant_category = restaurant_data.get("categories")
            return [
                [
                    product.get("isPopular", False),
                    product.get("name"),
                    product.get("description"),
                    product.get("priceNumber"),
                    product.get("realPrice"),
                    restaurant_name,
                    restaurant_status,
                    restaurant_category,
                ]
                for product_category in restaurant_data["corridors"]
                for product in product_category["products"]
            ]
        except Exception as error:
            self._metadata.num_errors += 1
            rest_name = restaurant.get_attribute("aria-label")
            LOGGER.error(
                f"Fallo al intentar extraer la información de los productos ofrecidos por el restaurante {rest_name}"
            )
            LOGGER.error(error)
            self._links_to_go.append((rest_name, restaurant_link, restaurant_status))
            return []

    def extract_data(self):
        """Extrae todos los productos que ofertan los restaurantes en Rappi"""
        LOGGER.info("Extrayendo los productos de rappi")
        self._driver.get("https://www.rappi.com.pe/restaurantes")
        del self._driver.requests

        # Extracción de información de los restaurantes más cercanos
        no_error = True
        start = 0
        while no_error:
            # Identificando el botón de ver más restaurantes
            try:
                button = self._wait.until(
                    EC.element_to_be_clickable(
                        (
                            By.XPATH,
                            "//button[@class='sc-hqyNC sc-jbKcbu bvSdOe primary wide']",
                        )
                    )
                )
            except Exception as error:
                no_error = False

            # Esperando a que se carguen los restaurantes
            time.sleep(random.uniform(2.0, 3.0))
            restaurants = self._wait.until(
                lambda x: x.find_elements(
                    By.XPATH,
                    "//div[@class='sc-c2b2dc55-4 bkatcD']/a",
                )
            )
            end = len(restaurants)
            # Interactuando con varios restaurantes a la vez
            future_restaurants = [
                THREAD.submit(self.extract_products, restaurants[index])
                for index in range(start, end)
            ]
            for future_restaurant in futures.wait(future_restaurants).done:
                self._products += future_restaurant.result()
            LOGGER.info(f"Cantidad de restaurantes totales recorridos: {end}")
            # Dar click al botón de ver más restaurantes
            try:
                del self._driver.requests
                button.click()
                start = end
            except:
                no_error = False

        LOGGER.info(
            f"Se extrajo la información de {len(self._restaurants)} restaurante(s)"
        )

        # Extracción de información de los restaurantes por categorías
        self._driver.get("https://www.rappi.com.pe/restaurantes")
        del self._driver.requests
        # Categorías de los restaurantes
        categories = self._driver.find_elements(
            By.XPATH, "//button[@class='sc-5d042f5c-1 iyWZJm']"
        )
        LOGGER.info(
            f"Se han detectado {len(categories)} tipos de restaurantes según su categoría"
        )
        for category in categories:
            LOGGER.info(f"Categoría: {category.text}")
            # Navegar a la parte superior de la página
            self._driver.execute_script(
                "window.scrollTo(document.body.scrollHeight, 0)"
            )
            time.sleep(random.uniform(1.0, 1.5))
            # Dar click a una categoría
            try:
                category.click()
                time.sleep(random.uniform(7.0, 8.0))
            except Exception as error:
                LOGGER.error("No se ha podido dar click a la categoría")
                LOGGER.error(f"Error de tipo {error.__class__}")
                continue
            # Dar click a la flecha de navegación de las categorías
            try:
                self._driver.find_element(By.CLASS_NAME, "sc-69ee8a42-2").click()
            except:
                pass
            # Identificar si la categoría cuenta con restaurantes
            try:
                # Identificar los restaurantes pertenecientes a la categoría seleccionada
                restaurants = self._wait.until(
                    lambda x: x.find_elements(
                        By.XPATH, "//div[@class='sc-c2b2dc55-4 bkatcD']/a"
                    )
                )
                # Filtrar los restaurantes con las que ya se cuente información
                restaurants = [
                    restaurant
                    for restaurant in restaurants
                    if restaurant.get_attribute("href") not in self._restaurants
                ]
                del self._driver.requests
                LOGGER.info(
                    f"Se va a extraer información de los productos de {len(restaurants)} restaurante(s)"
                )
            except:
                LOGGER.info("La categoría no cuenta con restaurantes")
                continue
            # Extraer la información de varios restaurantes a la vez
            future_restaurants = [
                THREAD.submit(self.extract_products, restaurant)
                for restaurant in restaurants
            ]
            for future_restaurant in futures.wait(future_restaurants).done:
                self._products += future_restaurant.result()
            del self._driver.requests

        # Eliminar valores duplicados:
        self._links_to_go = list(set(self._links_to_go))
        LOGGER.info(f"Se van a recorrer {len(self._links_to_go)} restaurantes")
        # Extrayendo la información de los restaurantes faltantes
        for rest_name, restaurant_link, restaurant_status in self._links_to_go:
            del self._driver.requests
            self.scrap_restaurante(rest_name, restaurant_link, restaurant_status)
        LOGGER.info("Extracción de datos completada satisfactoriamente")

    def process_data(self):
        """Proceso de limpieza de datos extraídos por el scraper"""
        try:
            LOGGER.info("Limpiando la data extraída por el scraper")
            self._dataset = pd.DataFrame(
                self._products,
                columns=[
                    "Popular",
                    "Producto",
                    "Descripcion",
                    "Precio con descuento",
                    "Precio sin descuento",
                    "Restaurante",
                    "Disponible",
                    "Categoria",
                ],
            )
            self._products = []
            self._dataset["Fecha"] = CURRENT_DATE.strftime("%Y-%m-%d")
            self._dataset.sort_values(
                ["Restaurante", "Producto", "Descripcion", "Popular"],
                inplace=True,
                ascending=[True, True, True, False],
            )
            self._dataset.drop_duplicates(
                ["Restaurante", "Producto", "Descripcion"], keep="first", inplace=True
            )
            self._dataset.replace({"": None}, inplace=True)
            self._dataset = self._dataset.astype(
                {
                    "Precio con descuento": float,
                    "Precio sin descuento": float,
                    "Popular": str,
                }
            )
            self._dataset["Precio con descuento"] = self._dataset[
                "Precio con descuento"
            ].apply(lambda x: round(x, 2))
            self._dataset.loc[
                self._dataset[
                    self._dataset["Precio con descuento"]
                    == self._dataset["Precio sin descuento"]
                ].index,
                "Precio con descuento",
            ] = None
            self._dataset[
                ["Precio con descuento", "Precio sin descuento"]
            ] = self._dataset[
                ["Precio con descuento", "Precio sin descuento"]
            ].applymap(
                "{:,.2f}".format
            )
            self._dataset[["Precio con descuento", "Precio sin descuento"]] = (
                self._dataset[["Precio con descuento", "Precio sin descuento"]]
                .replace({",": ";", "\.": ","}, regex=True)
                .replace({";": "."}, regex=True)
            )
            self._dataset["Disponible"].replace(
                "^Abre.+", "Restaurante cerrado", regex=True, inplace=True
            )
            self._dataset["Categoria"].replace(" -.+", "", regex=True, inplace=True)
            self._dataset["Restaurante"].replace(" -.+", "", regex=True, inplace=True)
            self._dataset["Popular"].replace(
                {"True": "popular", "False": ""}, inplace=True
            )
            LOGGER.info("Se ha limpiado la data satisfactoriamente")
        except Exception as error:
            LOGGER.error("Error al ejecutar el proceso completo de limpieza de datos")
            LOGGER.error(error)

    def save_data(self, filepath, filename, encoding="utf-8-sig"):
        """Guarda los datos o errores obtenidos durante la ejecución del scraper

        Args:
            filepath (str): Ruta del archivo
            filename (str): Nombre del archivo
            encoding (str): Codificación usada para guardar el archivo. Defaults to "utf-8-sig"
        """
        LOGGER.info("Guardando la data")
        # Comprobando que el dataset contenga información
        if len(self._dataset) == 0:
            LOGGER.info(
                f"El archivo de datos no se va a guardar por no tener información",
            )
            return

        self._metadata.quantity = len(self._dataset)
        # Generando la ruta donde se va a guardar la información
        filepath = os.path.join(filepath, CURRENT_DATE.strftime("%d-%m-%Y"))
        filename = (
            filename
            + "_"
            + CURRENT_DATE.strftime("%Y-%m-%d")
            + "_"
            + str(self._metadata.quantity)
            + ".csv"
        )

        # Verificando si la ruta donde se va a guardar la información existe
        if not os.path.exists(filepath):
            os.makedirs(filepath)

        self._dataset.to_csv(
            os.path.join(filepath, filename),
            sep=";",
            index=False,
            encoding=encoding,
        )
        LOGGER.info(
            f"El archivo de datos {filename} ha sido guardado correctamente en la ruta {os.path.join(ROOT_PATH, filepath)}",
        )

    def save_metadata(self, filename, sheet_name):
        """Guarda la información de la metadata generada durante la ejecución del scraper

        Args:
            filename (str): Nombre del archivo
            sheet_name (str): Nombre de la hoja de cálculo
        """
        LOGGER.info("Guardando la metadata")
        self._metadata.set_attributes_values()
        self._metadata.print_metadata_information()
        # Variable que indica si el encabezado existe o no en el archivo de excel
        header_exist = False

        # Verificando si el archivo existe o no
        if os.path.isfile(filename):
            wb_time = load_workbook(filename)
            # Comprobando si ya existe un sheet con el nombre indicado en la variable sheet_name
            if sheet_name not in [ws.title for ws in wb_time.worksheets]:
                # Creando un nuevo sheet
                wb_time.create_sheet(sheet_name)
            else:
                header_exist = True
        else:
            wb_time = Workbook()
            wb_time.worksheets[0].title = sheet_name

        # Seleccionar el sheet deseado donde se va a guardar la información
        worksheet = wb_time[sheet_name]

        # Comprobando si el encabezado existe o no
        if not header_exist:
            keys = [
                "Fecha",
                "Hora Inicio",
                "Hora Fin",
                "Cantidad",
                "Tiempo Ejecucion (min)",
                "Productos / Minuto",
                "Errores",
            ]
            worksheet.append(keys)

        values = list(self._metadata.__dict__.values())[2:]
        worksheet.append(values)
        wb_time.save(filename)
        wb_time.close()
        LOGGER.info(
            f"El archivo de la metadata del scraper {filename} ha sido guardado correctamente en la ruta {ROOT_PATH}",
        )

    def run(self):
        """Ejecuta el proceso completo de web scraping a Rappi"""
        self.login(FB_USERNAME, FB_PASSWORD)
        self.extract_data()
        self.process_data()
        self.save_data(DATA_FOLDER, DATA_FILENAME)
        self.save_metadata(METADATA_FILENAME, METADATA_SHEET_NAME)


def configure_log(
    log_folder, log_filename, log_file_mode="w", log_file_encoding="utf-8"
):
    """Función que configura los logs para rastrear al programa

    Args:
        log_folder (str): Carpeta donde se va a generar el archivo log
        log_filename (str): Nombre del archivo log a ser generado
        log_file_mode (str, optional): Modo de guardado del archivo. Defaults to "w".
        log_file_encoding (str, optional): Codificación usada para el archivo. Defaults to "utf-8".
    """
    # Generando la ruta donde se va a guardar los registros de ejecución
    log_path = os.path.join(log_folder, CURRENT_DATE.strftime("%d-%m-%Y"))
    log_filename = log_filename + "_" + CURRENT_DATE.strftime("%d%m%Y") + ".log"

    # Verificando si la ruta donde se va a guardar los registros de ejecución existe
    if not os.path.exists(log_path):
        os.makedirs(log_path)

    # Agregando los handlers al logger
    formatter = logging.Formatter("%(asctime)s - %(levelname)s - %(message)s")
    stream_handler = logging.StreamHandler()
    stream_handler.setFormatter(formatter)
    file_handler = logging.FileHandler(
        os.path.join(log_path, log_filename), log_file_mode, log_file_encoding
    )
    file_handler.setFormatter(formatter)
    LOGGER.handlers = [stream_handler, file_handler]
    LOGGER.propagate = False
    LOGGER.setLevel(logging.INFO)


def main():
    try:
        configure_log(LOG_FOLDER, LOG_FILENAME)
        scraper = ScraperRappiProducts()
        scraper.run()
        LOGGER.info("Programa finalizado")
    except Exception as error:
        Error(error).print_error_detail()
        LOGGER.error("Programa ejecutado con fallos")
    finally:
        # Liberar el archivo log
        logging.shutdown()


if __name__ == "__main__":
    main()
